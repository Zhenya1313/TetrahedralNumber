from openpyxl import load_workbook
from openpyxl import Workbook
import time
def isInt(value):
    try:
        int_value = int(value)
        return True
    except ValueError:
        return False
def generate_pyramid_numbers(count=None, start=None, end=None):
    pyramid_numbers = []

    if count is not None:
        for n in range(1, count + 1):
            number = n * (n + 1) * (n + 2) // 6
            pyramid_numbers.append(number)
    elif start is not None and end is not None:
        n = 1
        while True:
            number = n * (n + 1) * (n + 2) // 6
            if number > end:
                break
            if number >= start:
                pyramid_numbers.append(number)
            n += 1
    else:
        raise ValueError("Не коректні дані")

    return pyramid_numbers


def min_sum_number(target, pyramid_numbers=None, start_index=0, current_combination=None):
    pyramid_numbers = generate_pyramid_numbers(start=1,end=target)

    if current_combination is None:
        current_combination = []

    if target == 0:
        return current_combination

    if target < 0 or len(current_combination) > 5:
        return None

    min_combination = None

    for i in range(start_index, len(pyramid_numbers)):
        next_combo = current_combination + [pyramid_numbers[i]]
        result = min_sum_number(target - pyramid_numbers[i], pyramid_numbers, i, next_combo)

        if result is not None:
            if min_combination is None or len(result) < len(min_combination):
                min_combination = result

    return min_combination

def write_to_file(data, filename='data.xlsx'):
    try:
        wb = load_workbook(filename)
        ws = wb['data']

        ws.delete_rows(1, ws.max_row)

    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = 'data'

    for row in data:
        ws.append(row)

    wb.save(filename)
    wb.close()

def choose_options():
    print("Оберіть опцію:")
    print("1. Задати кількість чисел")
    print("2. Задати діапазон чисел")
    print("3. Перевірка числа (Сума пірамідальних чисел)")
    print("4. Перевірка массива")
    choice = input("Введіть опції: ").strip()

    if choice == '1':
        count = int(input("Введіть кількість чисел: ").strip())
        result = generate_pyramid_numbers(count=count)
    elif choice == '2':
        start = input("Введіть початкове число: ").strip()
        if isInt(start) == False:
            raise ValueError("Введіть ціле початкове число")

        end = input("Введіть кінцеве число: ").strip()
        if not isInt(end):
            raise ValueError("Введіть ціле кінцеве число")

        result = generate_pyramid_numbers(start=int(start), end=int(end))
    elif choice == '3':
        number = input("Введіть число: ").strip()
        if isInt(number) == False:
            raise ValueError("Введіть ціле число")

        result = min_sum_number(int(number))
    elif choice == '4':
        size = int(input("Введіть кількість чисел: ").strip())
        numbers = list(range(1, size + 1))
        start_time = time.time()
        if numbers:
            data_to_write = []
            for number in numbers:
                combination = min_sum_number(number)
                merged_array = [number, ''] + combination
                data_to_write.append(merged_array)

        end_time = time.time()

        execution_time = end_time - start_time
        write_to_file(data_to_write)

        result = f"Час виконання: {execution_time} секунд"
    else:
        print("Не коректно обрана опція")
        return

    print("Результат:", result)


choose_options()
